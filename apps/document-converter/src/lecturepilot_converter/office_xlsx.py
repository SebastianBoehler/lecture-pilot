from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MAX_CELLS_PER_SHEET = 10_000


class SpreadsheetConversionError(RuntimeError):
    pass


def xlsx_table_blocks(path: Path) -> list[dict]:
    formulas = load_workbook(path, read_only=True, data_only=False)
    values = load_workbook(path, read_only=True, data_only=True)
    try:
        return [
            _sheet_block(formulas[sheet_name], values[sheet_name])
            for sheet_name in formulas.sheetnames
        ]
    finally:
        formulas.close()
        values.close()


def _sheet_block(formula_sheet, value_sheet) -> dict:
    max_row = formula_sheet.max_row
    max_column = formula_sheet.max_column
    if max_row * max_column > MAX_CELLS_PER_SHEET:
        raise SpreadsheetConversionError("Spreadsheet sheet exceeds the normalized cell limit.")
    cells = []
    for row in formula_sheet.iter_rows():
        for cell in row:
            raw = cell.value
            if raw is None:
                continue
            formula = raw if cell.data_type == "f" else None
            value = value_sheet.cell(cell.row, cell.column).value if formula else raw
            cells.append(
                {
                    "row": cell.row,
                    "column": cell.column,
                    "value": value,
                    "formula": formula,
                }
            )
    end = f"{get_column_letter(max_column)}{max_row}"
    return {
        "kind": "table",
        "cells": cells,
        "locator": {"sheet": formula_sheet.title, "cell_range": f"A1:{end}"},
        "extraction": "native",
    }
